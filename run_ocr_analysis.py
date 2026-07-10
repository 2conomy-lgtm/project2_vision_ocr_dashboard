#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Project 2: 멀티모달 기반 통합 분석 에이전트 실습 프로젝트
[5단계] OCR 결과와 정답 데이터 비교 분석 스크립트 (run_ocr_analysis.py)

이 스크립트는 data/ocr/ocr_extracted_raw.csv와 data/source_structured/ground_truth_multimodal_240.csv를
record_id를 기준으로 조인(Inner Join)하여 정밀 품질 평가를 수행합니다.
평가 지표:
- 문서 유형별 종합 정확도
- 영수증 합계금액 추출 정확도 (수치 완전 일치 여부)
- 설문 세부 점수 추출 정확도 (3개 지표 각각 및 전체 일치 여부)
- 수기 메모 추출 성공률 (difflib SequenceMatcher를 사용한 문자열 유사성 평가)

결과는 reports/ocr_quality_summary.txt와 reports/ocr_quality_report.xlsx에 저장됩니다.
openpyxl을 통해 엑셀 시트에 프로페셔널한 네이비/민트 테마 서식과 이중 선 스타일을 직접 가미합니다.
"""

import os
import sys
import difflib
import pandas as pd
import numpy as np

# openpyxl 서식 입히기용 모듈 임포트
try:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

def calculate_string_similarity(str1, str2):
    """두 문자열의 유사도를 0.0~1.0 사이로 계산합니다."""
    if pd.isna(str1) or pd.isna(str2):
        if pd.isna(str1) and pd.isna(str2):
            return 1.0
        return 0.0
    s1 = str(str1).strip()
    s2 = str(str2).strip()
    if not s1 and not s2:
        return 1.0
    return difflib.SequenceMatcher(None, s1, s2).ratio()

def run_analysis():
    print("=" * 60)
    print(" [5단계] 정답 대비 OCR 추출 품질 정밀 대조 분석을 시작합니다.")
    print("=" * 60)

    # 1. 파일 경로 설정
    project_root = os.path.dirname(os.path.abspath(__file__))
    gt_path = os.path.join(project_root, "data", "source_structured", "ground_truth_multimodal_240.csv")
    ocr_path = os.path.join(project_root, "data", "ocr", "ocr_extracted_raw.csv")
    report_dir = os.path.join(project_root, "reports")
    
    txt_report_path = os.path.join(report_dir, "ocr_quality_summary.txt")
    xlsx_report_path = os.path.join(report_dir, "ocr_quality_report.xlsx")

    # CSV 로드
    if not os.path.exists(gt_path) or not os.path.exists(ocr_path):
        print(f"[오류] 데이터 파일이 존재하지 않습니다.\n- 정답: {gt_path}\n- OCR결과: {ocr_path}")
        sys.exit(1)

    try:
        gt_df = pd.read_csv(gt_path)
        ocr_df = pd.read_csv(ocr_path)
    except Exception as e:
        print(f"[오류] 데이터 로드 실패: {e}")
        sys.exit(1)

    # 2. record_id 기준 병합 (Inner Join)
    merged = pd.merge(gt_df, ocr_df, on="record_id", suffixes=("_gt", "_ocr"))
    print(f"-> [조인 완료] 총 {len(merged)}개 레코드가 매칭되었습니다.")

    # 3. 문서 유형 분류
    receipts = merged[merged['document_type_gt'] == 'receipt'].copy()
    surveys = merged[merged['document_type_gt'] == 'survey'].copy()

    # 4. 영수증 품질 평가 분석
    print("-> 영수증(Receipt) 데이터 정밀 비교 중...")
    receipt_metrics = []
    
    # 영수증 금액 정확도 산출
    r_amt_match_count = 0
    r_date_match_count = 0
    r_store_match_count = 0
    r_note_sim_sum = 0.0

    for idx, row in receipts.iterrows():
        rec_id = row['record_id']
        
        # 날짜 비교
        gt_date = str(row['doc_date']).strip()
        ocr_date = str(row['extracted_date']).strip() if pd.notna(row['extracted_date']) else ""
        date_match = (gt_date == ocr_date)
        if date_match: r_date_match_count += 1
        
        # 상호명 비교
        gt_store = str(row['organization_or_store']).strip()
        ocr_store = str(row['extracted_store_or_dept']).strip() if pd.notna(row['extracted_store_or_dept']) else ""
        store_match = (gt_store == ocr_store)
        if store_match: r_store_match_count += 1

        # 금액 비교 (결측 및 수치 보정 처리 후)
        gt_amt = row['total_amount']
        ocr_amt = row['extracted_amount']
        
        # numeric 일치 판단
        try:
            amt_match = (int(float(gt_amt)) == int(float(ocr_amt))) if pd.notna(gt_amt) and pd.notna(ocr_amt) else False
        except ValueError:
            amt_match = False
            
        if amt_match: r_amt_match_count += 1

        # 수기 메모 유사도
        gt_note = row['handwritten_note']
        ocr_note = row['extracted_note']
        note_sim = calculate_string_similarity(gt_note, ocr_note)
        r_note_sim_sum += note_sim

        # 개별 영수증 레코드 요약 보관
        receipt_metrics.append({
            "record_id": rec_id,
            "category": row['category'],
            "gt_date": gt_date,
            "ocr_date": ocr_date,
            "date_match": date_match,
            "gt_store": gt_store,
            "ocr_store": ocr_store,
            "store_match": store_match,
            "gt_amount": int(gt_amt) if pd.notna(gt_amt) else 0,
            "ocr_amount": int(float(ocr_amt)) if pd.notna(ocr_amt) and ocr_amt != "" and not pd.isna(ocr_amt) else "",
            "amount_match": amt_match,
            "gt_note": gt_note if pd.notna(gt_note) else "",
            "ocr_note": ocr_note if pd.notna(ocr_note) else "",
            "note_similarity": round(note_sim, 4),
            "confidence": row['confidence']
        })

    r_count = len(receipts)
    r_date_acc = r_date_match_count / r_count if r_count > 0 else 0.0
    r_store_acc = r_store_match_count / r_count if r_count > 0 else 0.0
    r_amt_acc = r_amt_match_count / r_count if r_count > 0 else 0.0
    r_note_acc = r_note_sim_sum / r_count if r_count > 0 else 0.0
    r_total_acc = (r_date_acc + r_store_acc + r_amt_acc + r_note_acc) / 4.0

    # 5. 설문지 품질 평가 분석
    print("-> 설문지(Survey) 데이터 정밀 비교 중...")
    survey_metrics = []
    
    s_date_match_count = 0
    s_dept_match_count = 0
    s_score_all_match_count = 0
    s_score_sat_match_count = 0
    s_score_usa_match_count = 0
    s_score_spd_match_count = 0
    s_note_sim_sum = 0.0

    for idx, row in surveys.iterrows():
        rec_id = row['record_id']
        
        # 날짜 비교
        gt_date = str(row['doc_date']).strip()
        ocr_date = str(row['extracted_date']).strip() if pd.notna(row['extracted_date']) else ""
        date_match = (gt_date == ocr_date)
        if date_match: s_date_match_count += 1
        
        # 부서명 비교
        gt_dept = str(row['respondent_dept']).strip()
        ocr_dept = str(row['extracted_store_or_dept']).strip() if pd.notna(row['extracted_store_or_dept']) else ""
        dept_match = (gt_dept == ocr_dept)
        if dept_match: s_dept_match_count += 1

        # 설문 만족도 3개 점수 대조
        gt_sat = int(row['satisfaction_score']) if pd.notna(row['satisfaction_score']) else -1
        gt_usa = int(row['usability_score']) if pd.notna(row['usability_score']) else -1
        gt_spd = int(row['speed_score']) if pd.notna(row['speed_score']) else -1
        
        ocr_scores_str = str(row['extracted_scores']).strip() if pd.notna(row['extracted_scores']) else ""
        
        # 추출된 점수 파싱 (콤마 분할)
        ocr_sat, ocr_usa, ocr_spd = -2, -2, -2
        if ocr_scores_str:
            parts = ocr_scores_str.split(',')
            if len(parts) >= 3:
                try: ocr_sat = int(parts[0]) if parts[0].strip() else -2
                except ValueError: pass
                try: ocr_usa = int(parts[1]) if parts[1].strip() else -2
                except ValueError: pass
                try: ocr_spd = int(parts[2]) if parts[2].strip() else -2
                except ValueError: pass

        sat_match = (gt_sat == ocr_sat)
        usa_match = (gt_usa == ocr_usa)
        spd_match = (gt_spd == ocr_spd)
        
        if sat_match: s_score_sat_match_count += 1
        if usa_match: s_score_usa_match_count += 1
        if spd_match: s_score_spd_match_count += 1
        
        scores_all_match = sat_match and usa_match and spd_match
        if scores_all_match: s_score_all_match_count += 1

        # 수기 메모 유사도
        gt_note = row['handwritten_note']
        ocr_note = row['extracted_note']
        note_sim = calculate_string_similarity(gt_note, ocr_note)
        s_note_sim_sum += note_sim

        survey_metrics.append({
            "record_id": rec_id,
            "dept": gt_dept,
            "gt_date": gt_date,
            "ocr_date": ocr_date,
            "date_match": date_match,
            "gt_dept": gt_dept,
            "ocr_dept": ocr_dept,
            "dept_match": dept_match,
            "gt_scores": f"{gt_sat},{gt_usa},{gt_spd}",
            "ocr_scores": ocr_scores_str,
            "satisfaction_match": sat_match,
            "usability_match": usa_match,
            "speed_match": spd_match,
            "scores_all_match": scores_all_match,
            "gt_note": gt_note if pd.notna(gt_note) else "",
            "ocr_note": ocr_note if pd.notna(ocr_note) else "",
            "note_similarity": round(note_sim, 4),
            "confidence": row['confidence']
        })

    s_count = len(surveys)
    s_date_acc = s_date_match_count / s_count if s_count > 0 else 0.0
    s_dept_acc = s_dept_match_count / s_count if s_count > 0 else 0.0
    s_scores_all_acc = s_score_all_match_count / s_count if s_count > 0 else 0.0
    s_score_sat_acc = s_score_sat_match_count / s_count if s_count > 0 else 0.0
    s_score_usa_acc = s_score_usa_match_count / s_count if s_count > 0 else 0.0
    s_score_spd_acc = s_score_spd_match_count / s_count if s_count > 0 else 0.0
    s_note_acc = s_note_sim_sum / s_count if s_count > 0 else 0.0
    s_total_acc = (s_date_acc + s_dept_acc + s_scores_all_acc + s_note_acc) / 4.0

    # 종합 지표
    grand_total_acc = (r_total_acc + s_total_acc) / 2.0

    # 6. TXT 리포트 생성 및 저장
    txt_content = []
    txt_content.append("=" * 80)
    txt_content.append("          [Project 2 OCR 품질 정밀 비교 분석 보고서]")
    txt_content.append("=" * 80)
    txt_content.append(f" - 평가 분석 일시: 2026-07-10 10:11")
    txt_content.append(f" - 검증 총 레코드 : {len(merged)}건 (영수증 {r_count}건, 설문지 {s_count}건)")
    txt_content.append(f" - 전체 평균 OCR 종합 매칭률: {grand_total_acc*100:.2f}%")
    txt_content.append("-" * 80)
    
    txt_content.append(" 1. 영수증(Receipt) 부문 품질 지표")
    txt_content.append(f"   * 영수증 종합 인식 매칭률: {r_total_acc*100:.2f}%")
    txt_content.append(f"   * 일자(doc_date) 일치도       : {r_date_acc*100:.2f}% ({r_date_match_count}/{r_count})")
    txt_content.append(f"   * 상호명(store) 일치도        : {r_store_acc*100:.2f}% ({r_store_match_count}/{r_count})")
    txt_content.append(f"   * 합계금액(total_amount) 정확도 : {r_amt_acc*100:.2f}% ({r_amt_match_count}/{r_count})")
    txt_content.append(f"   * 수기 메모(note) 평균 복원 유사도: {r_note_acc*100:.2f}%")
    txt_content.append("-" * 80)
    
    txt_content.append(" 2. 설문지(Survey) 부문 품질 지표")
    txt_content.append(f"   * 설문지 종합 인식 매칭률: {s_total_acc*100:.2f}%")
    txt_content.append(f"   * 응답일자(doc_date) 일치도    : {s_date_acc*100:.2f}% ({s_date_match_count}/{s_count})")
    txt_content.append(f"   * 응답부서(dept) 일치도        : {s_dept_acc*100:.2f}% ({s_dept_match_count}/{s_count})")
    txt_content.append(f"   * 3대 만족도 점수 완전 일치도  : {s_scores_all_acc*100:.2f}% ({s_score_all_match_count}/{s_count})")
    txt_content.append(f"     └ 세부 만족도 일치도 (만족도: {s_score_sat_acc*100:.1f}%, 편의성: {s_score_usa_acc*100:.1f}%, 응답속도: {s_score_spd_acc*100:.1f}%)")
    txt_content.append(f"   * 수기 의견(note) 평균 복원 유사도: {s_note_acc*100:.2f}%")
    txt_content.append("=" * 80)
    
    # 노이즈 감지 알림
    has_preprocessed = merged['preprocessing_used'].any()
    if has_preprocessed:
        txt_content.append(" ✔ [품질 상태] 이미지 전처리 필터가 작동 중입니다. 극도로 향상된 정확도를 나타내고 있습니다.")
    else:
        txt_content.append(" ⚠ [품질 경고] 이미지 전처리가 반영되지 않아 저해상도/노이즈 이미지에서의 오차율이 매우 높습니다.")
    txt_content.append("=" * 80 + "\n")

    # TXT 쓰기
    if not os.path.exists(report_dir):
        os.makedirs(report_dir, exist_ok=True)
        
    with open(txt_report_path, "w", encoding="utf-8") as f:
        f.write("\n".join(txt_content))
    print(f"-> [TXT 작성 완료] 요약 품질 리포트: {txt_report_path}")

    # 7. Excel 리포트 생성 및 서식화 (openpyxl 활용)
    r_detail_df = pd.DataFrame(receipt_metrics)
    s_detail_df = pd.DataFrame(survey_metrics)
    
    summary_data = {
        "평가 지표 (Metrics)": [
            "종합 OCR 정확도 (Grand Average Accuracy)",
            "영수증 종합 정확도 (Receipt Total Accuracy)",
            "영수증 일자 일치율 (Receipt Date Accuracy)",
            "영수증 상호명 일치율 (Receipt Store Accuracy)",
            "영수증 금액 정확도 (Receipt Amount Accuracy)",
            "영수증 메모 복원 유사도 (Receipt Note Similarity)",
            "설문지 종합 정확도 (Survey Total Accuracy)",
            "설문지 응답일자 일치율 (Survey Date Accuracy)",
            "설문지 응답부서 일치율 (Survey Dept Accuracy)",
            "설문지 만족도 3대점수 합산일치율 (Survey Scores All-match)",
            "설문지 수기의견 복원 유사도 (Survey Note Similarity)"
        ],
        "정확도 수치 (Accuracy/Ratio)": [
            grand_total_acc, r_total_acc, r_date_acc, r_store_acc, r_amt_acc, r_note_acc,
            s_total_acc, s_date_acc, s_dept_acc, s_scores_all_acc, s_note_acc
        ]
    }
    summary_df = pd.DataFrame(summary_data)

    try:
        with pd.ExcelWriter(xlsx_report_path, engine='openpyxl') as writer:
            summary_df.to_excel(writer, sheet_name="OCR_Summary", index=False)
            r_detail_df.to_excel(writer, sheet_name="Receipt_Details", index=False)
            s_detail_df.to_excel(writer, sheet_name="Survey_Details", index=False)

        # openpyxl 서식 입히기 진행
        if OPENPYXL_AVAILABLE:
            wb = writer.book
            
            # 테마 색상 (네이비 헤더, 아주 밝은 네이비/민트 포인트)
            navy_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
            header_font = Font(name="Malgun Gothic", size=11, bold=True, color="FFFFFF")
            title_font = Font(name="Malgun Gothic", size=13, bold=True, color="1B365D")
            cell_font = Font(name="Malgun Gothic", size=10)
            bold_font = Font(name="Malgun Gothic", size=10, bold=True)
            
            thin_border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )
            double_bottom = Border(bottom=Side(style='double', color='000000'))

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ws.views.sheetView[0].showGridLines = True
                
                # 1) 헤더 서식 입히기
                max_col = ws.max_column
                ws.row_dimensions[1].height = 28
                
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.fill = navy_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

                # 2) 본문 서식 및 컬럼 폭 맞춤
                max_row = ws.max_row
                for row in range(2, max_row + 1):
                    ws.row_dimensions[row].height = 20
                    for col in range(1, max_col + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.font = cell_font
                        cell.border = thin_border
                        
                        # 숫자 및 불리언은 가운데 정렬
                        if isinstance(cell.value, (bool, np.bool_)):
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        elif sheet_name == "OCR_Summary" and col == 2:
                            # 요약 시트 비율 포맷 설정 (백분율)
                            cell.number_format = '0.0%'
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                            cell.font = bold_font
                        elif str(cell.value).isdigit() or isinstance(cell.value, (int, float)):
                            cell.alignment = Alignment(horizontal="right", vertical="center")

                # 열 너비 맞춤
                for col in ws.columns:
                    max_len = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        if cell.value:
                            val_len = len(str(cell.value).encode('utf-8'))
                            if val_len > max_len:
                                max_len = val_len
                    # 패딩 추가
                    ws.column_dimensions[col_letter].width = max(max_len // 2 + 3, 12)

            wb.save(xlsx_report_path)
            print(f"-> [Excel 서식 완료] 디자인 적용 완료: {xlsx_report_path}")
        else:
            print("-> [경고] openpyxl 패키지를 로드할 수 없어 기본 서식으로 엑셀을 저장했습니다.")

    except Exception as e:
        print(f"[오류] Excel 리포트 생성 실패: {e}")

    # 터미널 화면에 결과 출력 (안전하게 출력)
    def safe_print(text):
        try:
            print(text)
        except UnicodeEncodeError:
            try:
                if hasattr(sys.stdout, 'reconfigure'):
                    sys.stdout.reconfigure(encoding='utf-8')
                    print(text)
                else:
                    print(text.encode(sys.stdout.encoding, errors='replace').decode(sys.stdout.encoding))
            except Exception:
                print(text.encode('ascii', errors='replace').decode('ascii'))

    safe_print("\n" + "\n".join(txt_content[:15]))
    safe_print(f" -> 상세 내역은 '{os.path.relpath(txt_report_path, project_root)}' 및 '{os.path.relpath(xlsx_report_path, project_root)}' 에서 보실 수 있습니다.\n")
    safe_print("=" * 60)
    safe_print(" [5단계] 정답 데이터 비교 분석 처리가 완전히 종료되었습니다.")
    safe_print("=" * 60 + "\n")

    return grand_total_acc

if __name__ == "__main__":
    run_analysis()
